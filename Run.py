from openpyxl import Workbook
from datetime import datetime
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import DateAxis
from openpyxl.chart import BarChart, Series, Reference
from copy import deepcopy
from openpyxl.chart import (
    PieChart,
    LineChart,
    ProjectedPieChart,
    Reference
)


def function (siteName,includeNoneed):
    wb = load_workbook(filename = 'SiteStatusProject.xlsx')
    sheet1 = wb['SiteStatusProject']

    #iterate over data-

    maxRow = sheet1.max_row + 1
    total =0
    done =0

    for row in range(2, maxRow):
        if(sheet1['C' +str(row)].value==siteName or siteName=="All sites"):
            if ((sheet1['E' + str(row)].value == "No Need" and includeNoneed == True) or (sheet1['E' + str(row)].value != "No Need" )):
                total+=1
                if(sheet1['B' +str(row)].value=="Done"):
                    done += 1


    rows = [
        ('Status', 'Amount'),
        ("Planned", total),
        ("Actual", done),
    ]

    print (rows)

    print(siteName)
    print("\n")

    ws2 = wb.create_sheet(title= siteName)

    for row in rows:
        ws2.append(row)


    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    if(includeNoneed==True):
        chart1.title = "Plan vs. Actual"
    else:
        chart1.title = "Plan vs. Actual - Without 'No Need' "

    chart1.y_axis.title = "Flight number"
    chart1.x_axis.title = "Total"
    data = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row, max_col=ws2.max_column)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=3)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.dataLabels = DataLabelList()
    chart1.dataLabels.showVal = True
    chart1.shape = 4



    ws2.add_chart(chart1, "E2")
    wb.save('SiteStatusProject.xlsx')

    maxRow = sheet1.max_row + 1
    total = 0
    done = 0
    dict= {}
    rows =[]
    for row in range(2, maxRow):
        if (sheet1['C' + str(row)].value == siteName or siteName == "All sites"):
            if ((sheet1['E' + str(row)].value == "No Need" and includeNoneed == True) or (sheet1['E' + str(row)].value != "No Need" )):
                    print(sheet1['D' + str(row)].value.strftime('%d/%m/%Y'))
                    if (dict.__contains__(sheet1['D' + str(row)].value.strftime('%d/%m/%Y')) == False):
                        dict[sheet1['D' + str(row)].value.strftime('%d/%m/%Y')] = {}
                        total += 1
                        dict[sheet1['D' + str(row)].value.strftime('%d/%m/%Y')]['total'] = total
                        dict[sheet1['D' + str(row)].value.strftime('%d/%m/%Y')]['done'] = done
                        if (sheet1['B' + str(row)].value == "Done"):
                            done += 1
                            dict[sheet1['D' + str(row)].value.strftime('%d/%m/%Y')]['done'] = done
                    else:
                        dict[sheet1['D' + str(row)].value.strftime('%d/%m/%Y')]['total'] += 1
                        if (sheet1['B' + str(row)].value == "Done"):
                            dict[sheet1['D' + str(row)].value.strftime('%d/%m/%Y')]['done'] += 1
        total = 0
        done = 0

    #inset into rows
    row=[]
    row.append("date")
    row.append("Planned")
    row.append("Actual")
    rows.append(row)


    for k in dict:
        print(k)
        k=datetime.strptime(k, '%d/%m/%Y')


    ordered_data = sorted(dict.items(), key=lambda x: datetime.strptime(x[0],'%d/%m/%Y'), reverse=False)
    print("DATA")
    #print(ordered_data)

    for data in ordered_data:
            row=[]
            row.append(data[0])
            row.append(data[1]['total'])
            row.append(data[1]['done'])
            rows.append(row)


    print(rows)


    for row in rows:
        ws2.append(row)

    data = Reference(ws2, min_col=2, min_row=4, max_col=3, max_row=ws2.max_row+1)

    # Chart with date axis
    c2 = LineChart()
    if (includeNoneed == True):
        c2.title = "Plan vs. Actual"
    else:
        c2.title = "Plan vs. Actual - Without 'No Need' "

    c2.style = 12
    c2.y_axis.title = "Flight number"
    c2.y_axis.crossAx = 500
    c2.x_axis = DateAxis(crossAx=100)
    c2.x_axis.number_format = 'd-mmm'
    c2.x_axis.majorTimeUnit = "days"
    c2.dataLabels = DataLabelList()
    c2.dataLabels.showVal = True
    c2.add_data(data, titles_from_data=True)
    dates = Reference(ws2, min_col=1, min_row=5, max_row=ws2.max_row+1)
    c2.set_categories(dates)

    ws2.add_chart(c2, "E18")
    wb.save('SiteStatusProject.xlsx')

    #iterate over data-

def failReasonNoNeed(siteName ,includeNoneed):
    wb = load_workbook(filename='SiteStatusProject.xlsx')
    sheet1 = wb['SiteStatusProject']
    maxRow = sheet1.max_row + 1

    fail =0
    cancel =0
    postponed =0
    precipitation=0
    wind=0
    lowBattery=0
    gpsLoss=0
    gpsBlock=0
    els=0
    systemMal=0
    fog=0
    siteTeam=0
    client=0
    noNeed=0
    systemFailure=0
    external=0
    corporate=0
    other=0
    siteTeamC=0
    clientC=0
    noNeedC=0
    systemFailureC=0
    externalC=0
    corporateC=0
    otherC =0
    for row in range(2, maxRow):
        if(sheet1['C' +str(row)].value==siteName or siteName=="All sites"):
            if((sheet1['E' + str(row)].value == "No Need" and includeNoneed == True) or (sheet1['E' + str(row)].value != "No Need") ):
                if(sheet1['B' +str(row)].value=="Fail"):
                    fail+=1
                    if (sheet1['F' + str(row)].value == "Precipitation"):
                        precipitation += 1
                    elif (sheet1['F' + str(row)].value == "Wind"):
                        wind += 1
                    elif (sheet1['F' + str(row)].value == "Low battery power"):
                        lowBattery += 1
                    elif (sheet1['F' + str(row)].value == "GPS loss"):
                        gpsLoss += 1
                    elif (sheet1['F' + str(row)].value == "GPS block"):
                        gpsBlock += 1
                    elif (sheet1['F' + str(row)].value == "ELS landing"):
                        els += 1
                    elif (sheet1['F' + str(row)].value == "System malfunction "):
                        systemMal += 1
                    elif (sheet1['F' + str(row)].value == "Fog"):
                        fog += 1
                elif(sheet1['B' +str(row)].value=="Cancelled"):
                    cancel+=1
                    if (sheet1['E' + str(row)].value == "Site Team Error"):
                        siteTeamC += 1
                    elif (sheet1['E' + str(row)].value == "Client Retaliated Reason"):
                        clientC += 1
                    elif (sheet1['E' + str(row)].value == "No Need"):
                        noNeedC += 1
                    elif (sheet1['E' + str(row)].value == "System Failure"):
                        systemFailureC += 1
                    elif (sheet1['E' + str(row)].value == "External Reason"):
                        externalC += 1
                    elif (sheet1['E' + str(row)].value == "Corporate Support"):
                        corporateC += 1
                    elif (sheet1['E' + str(row)].value == "Other"):
                        otherC += 1
                elif(sheet1['B' +str(row)].value=="Postponed"):
                    postponed+=1
                    if (sheet1['E' + str(row)].value == "Site Team Error"):
                        siteTeam += 1
                    elif (sheet1['E' + str(row)].value == "Client Retaliated Reason"):
                        client += 1
                    elif (sheet1['E' + str(row)].value == "No Need"):
                        noNeed += 1
                    elif (sheet1['E' + str(row)].value == "System Failure"):
                        systemFailure += 1
                    elif (sheet1['E' + str(row)].value == "External Reason"):
                        external += 1
                    elif (sheet1['E' + str(row)].value == "Corporate Support"):
                        corporate += 1
                    elif (sheet1['E' + str(row)].value == "Other"):
                        other += 1


    print(siteName)
    print("fail- " ,fail)
    print("cancel- ",cancel)
    print("postponed- ",postponed)
    print("\n")

    ws2 = wb.create_sheet(title=siteName)

    data = [
        ['fail reason', 'Count'],
        ['fail', int(fail)],
        ['cancel', int(cancel)],
        ['postponed', int(postponed)],
    ]

    dataFail = [
        ['fail reason', 'Count'],
        ['Precipitation', int(precipitation)],
        ['wind', int(wind)],
        ['Low battery power', int(lowBattery)],
        ['GPS loss', int(gpsLoss)],
        ['GPS block', int(gpsBlock)],
        ['ELS landing', int(els)],
        ['System malfunction', int(systemMal)],
        ['Fog', int(fog)],

    ]

    dataPostponed = [
        ['postponed reason', 'Count'],
        ['Site Team Error', int(siteTeam)],
        ['Client Retaliated Reason', int(client)],
        ['No Need', int(noNeed)],
        ['System Failure', int(systemFailure)],
        ['External Reason', int(external)],
        ['Corporate Support', int(corporate)],
        ['Other', int(other)],
    ]
    dataCancelled = [
        ['cancel reason', 'Count'],
        ['Site Team Error', int(siteTeamC)],
        ['Client Retaliated Reason', int(clientC)],
        ['No Need', int(noNeedC)],
        ['System Failure', int(systemFailureC)],
        ['External Reason', int(externalC)],
        ['Corporate Support', int(corporateC)],
        ['Other', int(otherC)],
    ]

    #total data

    for row in data:
        ws2.append(row)

    pie = PieChart()
    labels = Reference(ws2, min_col=1, min_row=2, max_row=4)
    data = Reference(ws2, min_col=2, min_row=1, max_row=4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    if (includeNoneed == True):
        pie.title = "Mission fail / cancel / postponed"
    else:
        pie.title = "Mission fail / cancel / postponed - Without 'No Need' "

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True

    ws2.add_chart(pie, "D3")

    #fail data

    for row in dataFail:
        ws2.append(row)

    pie2 = PieChart()
    labels = Reference(ws2, min_col=1, min_row=6, max_row=12)
    data = Reference(ws2, min_col=2, min_row=5, max_row=12)
    pie2.add_data(data, titles_from_data=True)
    pie2.set_categories(labels)
    pie2.title = "Fail reason"
    pie2.dataLabels = DataLabelList()
    pie2.dataLabels.showPercent = True

    ws2.add_chart(pie2, "D19")

    #postponed reason
    for row in dataPostponed:
        ws2.append(row)

    pie3 = PieChart()
    labels = Reference(ws2, min_col=1, min_row=15, max_row=20)
    data = Reference(ws2, min_col=2, min_row=14, max_row=20)
    pie3.add_data(data, titles_from_data=True)
    pie3.set_categories(labels)
    pie3.title = "Postponed reason"
    pie3.dataLabels = DataLabelList()
    pie3.dataLabels.showPercent = True

    ws2.add_chart(pie3, "D35")

    #Cancel reason
    for row in dataCancelled:
        ws2.append(row)

    pie4 = PieChart()
    labels = Reference(ws2, min_col=1, min_row=23, max_row=28)
    data = Reference(ws2, min_col=2, min_row=22, max_row=28)
    pie4.add_data(data, titles_from_data=True)
    pie4.set_categories(labels)
    pie4.title = "Cancel reason"
    pie4.dataLabels = DataLabelList()
    pie4.dataLabels.showPercent = True

    ws2.add_chart(pie4, "D51")

    wb.save('SiteStatusProject.xlsx')



list = ["All sites","Haifa Bay Port" ,"FAB 28","S32 Mine","S32 Worsley alumina" ,"BHP Area C" ,"Minera Centinela","BHP San Manuel","Vale NC1"]
for siteName in list:
    function(siteName, True)
    function(siteName, False)
    failReasonNoNeed(siteName, True)
    failReasonNoNeed(siteName, False)

