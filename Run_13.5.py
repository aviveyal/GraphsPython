from openpyxl import Workbook
from datetime import datetime
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import DateAxis
from openpyxl.chart import BarChart, Series, Reference
from copy import deepcopy
from openpyxl.chart import (
    PieChart,
    LineChart,
    ProjectedPieChart,
    Reference
)


def function (siteName,includeNoneed):
    wb = load_workbook(filename = 'SiteStatusProject.xlsx')
    sheet1 = wb['SiteStatusProject']

    #iterate over data-

    maxRow = sheet1.max_row + 1
    total =0
    done =0

    for row in range(2, maxRow):
        if(sheet1['C' +str(row)].value==siteName or siteName=="All sites"):
            if ((sheet1['E' + str(row)].value == "No Need" and includeNoneed == True) or (sheet1['E' + str(row)].value != "No Need" )):
                total+=1
                if(sheet1['B' +str(row)].value=="Done"):
                    done += 1


    rows = [
        ('Status', 'Amount'),
        ("Planned", total),
        ("Actual", done),
    ]

    print (rows)

    print(siteName)
    print("\n")

    if (includeNoneed==False):
        ws2 = wb.create_sheet(title= siteName+"-WNN")
    else:
        ws2 = wb.create_sheet(title=siteName)

    for row in rows:
        ws2.append(row)


    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10

    chart1.title = "Plan vs. Actual"



    chart1.y_axis.title = "Flight number"
    chart1.x_axis.title = "Total"
    data = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row, max_col=ws2.max_column)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=3)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.dataLabels = DataLabelList()
    chart1.dataLabels.showVal = True
    chart1.shape = 4



    ws2.add_chart(chart1, "E2")
    wb.save('SiteStatusProject.xlsx')

    maxRow = sheet1.max_row + 1
    total = 0
    done = 0
    dict= {}
    rows =[]
    for row in range(2, maxRow):
        if (sheet1['C' + str(row)].value == siteName or siteName == "All sites"):
            if ((sheet1['E' + str(row)].value == "No Need" and includeNoneed == True) or (sheet1['E' + str(row)].value != "No Need" )):
                    print(sheet1['D' + str(row)].value.strftime('%d/%m'))
                    if (dict.__contains__(sheet1['D' + str(row)].value.strftime('%d/%m')) == False):
                        dict[sheet1['D' + str(row)].value.strftime('%d/%m')] = {}
                        total += 1
                        dict[sheet1['D' + str(row)].value.strftime('%d/%m')]['total'] = total
                        dict[sheet1['D' + str(row)].value.strftime('%d/%m')]['done'] = done
                        if (sheet1['B' + str(row)].value == "Done"):
                            done += 1
                            dict[sheet1['D' + str(row)].value.strftime('%d/%m')]['done'] = done
                    else:
                        dict[sheet1['D' + str(row)].value.strftime('%d/%m')]['total'] += 1
                        if (sheet1['B' + str(row)].value == "Done"):
                            dict[sheet1['D' + str(row)].value.strftime('%d/%m')]['done'] += 1
        total = 0
        done = 0

    #inset into rows
    row=[]
    row.append("date")
    row.append("Planned")
    row.append("Actual")
    rows.append(row)


    for k in dict:
        print(k)
        k=datetime.strptime(k, '%d/%m')


    ordered_data = sorted(dict.items(), key=lambda x: datetime.strptime(x[0],'%d/%m'), reverse=False)
    print("DATA")
    #print(ordered_data)

    for data in ordered_data:
            row=[]
            row.append(data[0])
            row.append(data[1]['total'])
            row.append(data[1]['done'])
            rows.append(row)


    print(rows)


    for row in rows:
        ws2.append(row)

    data = Reference(ws2, min_col=2, min_row=4, max_col=3, max_row=ws2.max_row+1)

    # Chart with date axis
    c2 = LineChart()

    c2.title = "Plan vs. Actual"


    c2.style = 12
    c2.y_axis.title = "Flight number"
    c2.y_axis.crossAx = 500
    c2.x_axis = DateAxis(crossAx=100)
    c2.x_axis.number_format = 'd-mmm'
    c2.x_axis.majorTimeUnit = "days"
    c2.dataLabels = DataLabelList()
    c2.dataLabels.showVal = True
    c2.add_data(data, titles_from_data=True)
    dates = Reference(ws2, min_col=1, min_row=5, max_row=ws2.max_row+1)
    c2.set_categories(dates)

    ws2.add_chart(c2, "E18")
    wb.save('SiteStatusProject.xlsx')

    #iterate over data-

def failReasonNoNeed(siteName ,includeNoneed):
    wb = load_workbook(filename='SiteStatusProject.xlsx')
    sheet1 = wb['SiteStatusProject']
    maxRow = sheet1.max_row + 1

    fail =0
    cancel =0
    done=0

    for row in range(2, maxRow):
        if(sheet1['C' +str(row)].value==siteName or siteName=="All sites"):
            if((sheet1['E' + str(row)].value == "No Need" and includeNoneed == True) or (sheet1['E' + str(row)].value != "No Need") ):
                if (sheet1['B' + str(row)].value == "Done"):
                    done+=1
                elif(sheet1['B' +str(row)].value=="Fail"):
                    fail+=1
                elif(sheet1['B' +str(row)].value=="Cancelled"):
                    cancel+=1


    print(siteName)
    print("fail- " ,fail)
    print("cancel- ",cancel)
    print("done- ", done)
    print("\n")

    if (includeNoneed==False):
        ws2 = wb.create_sheet(title= siteName+"-WNN")
    else:
        ws2 = wb.create_sheet(title=siteName)

    data = [
        ['Status', 'Count'],
        ['fail', int(fail)],
        ['cancel', int(cancel)],
        ['success', int(done)],
    ]


    #total data

    for row in data:
        ws2.append(row)

    pie = PieChart()
    labels = Reference(ws2, min_col=1, min_row=2, max_row=4)
    data = Reference(ws2, min_col=2, min_row=1, max_row=4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)


    pie.title = "Flight Mission status"


    pie.dataLabels = DataLabelList()
    pie.dataLabels.showVal = True

    ws2.add_chart(pie, "D3")


    wb.save('SiteStatusProject.xlsx')



list = ["All sites","Haifa Bay Port" ,"FAB 28","S32 Mine","S32 Worsley alumina" ,"BHP Area C" ,"Minera Centinela","BHP San Manuel","Vale NC1"]
for siteName in list:
    # function(siteName, True)
    # failReasonNoNeed(siteName, True)
    function(siteName, False)
    failReasonNoNeed(siteName, False)

